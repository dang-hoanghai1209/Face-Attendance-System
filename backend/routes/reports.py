import io
import csv
import json
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from database import get_db
from models.attendance import Attendance
from models.security_alert import SecurityAlert
from models.student import Student
from services.auth_service import get_current_user, require_admin
from services import report_service


_FONT_CANDIDATES = [
    (
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/arialbd.ttf",
    ),
    (
        "C:/Windows/Fonts/tahoma.ttf",
        "C:/Windows/Fonts/tahomabd.ttf",
    ),
    (
        "C:/Windows/Fonts/segoeui.ttf",
        "C:/Windows/Fonts/segoeuib.ttf",
    ),
    (
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    ),
]


def _register_pdf_fonts():
    for regular_path, bold_path in _FONT_CANDIDATES:
        regular = Path(regular_path)
        bold = Path(bold_path)
        if not regular.exists():
            continue

        try:
            pdfmetrics.registerFont(TTFont("ReportUnicode", str(regular)))
            pdfmetrics.registerFont(TTFont("ReportUnicodeBold", str(bold if bold.exists() else regular)))
            return "ReportUnicode", "ReportUnicodeBold"
        except Exception:
            continue

    return "Helvetica", "Helvetica-Bold"


_PDF_FONT, _PDF_FONT_BOLD = _register_pdf_fonts()

PAGE_W, PAGE_H = A4
MARGIN = 40
BASE_DIR = Path(__file__).resolve().parents[1]
EVALUATION_REPORTS_DIR = BASE_DIR / "reports"
EVALUATION_METRICS_CSV = EVALUATION_REPORTS_DIR / "evaluation_metrics.csv"
EVALUATION_DETAILS_CSV = EVALUATION_REPORTS_DIR / "evaluation_details.csv"
MODEL_TEST_LOG_CSV = EVALUATION_REPORTS_DIR / "model_test_log.csv"
STATUS_LABELS = {
    "present": "Có mặt",
    "late": "Đi trễ",
    "absent": "Vắng mặt",
}
SUMMARY_COLUMNS = {
    "student_code": "Mã SV",
    "full_name": "Họ tên",
    "class_name": "Lớp",
    "present": "Có mặt",
    "late": "Đi trễ",
    "absent": "Vắng mặt",
    "attended": "Tổng có mặt",
    "total_sessions": "Tổng buổi",
    "rate": "Tỷ lệ chuyên cần",
    "warning": "Cảnh báo",
}
SESSION_COLUMNS = {
    "student_code": "Mã SV",
    "full_name": "Họ tên",
    "class_name": "Lớp",
    "subject": "Môn học",
    "session_date": "Ngày học",
    "start_time": "Giờ bắt đầu",
    "end_time": "Giờ kết thúc",
    "status": "Trạng thái",
    "check_in_at": "Vào lớp",
    "check_out_at": "Ra về",
    "check_in_conf": "Độ tin cậy khi vào",
    "check_out_conf": "Độ tin cậy khi ra",
    "note": "Ghi chú",
}

ATTENDANCE_CSV_COLUMNS = [
    "session_id",
    "class_name",
    "student_code",
    "full_name",
    "attendance_status",
    "check_in_at",
    "check_out_at",
    "check_in_conf",
    "check_out_conf",
    "liveness_passed",
    "gps_lat",
    "gps_lng",
    "gps_accuracy",
    "distance_meters",
    "check_in_img",
    "scan_count",
    "last_scan_at",
    "note",
    "created_at",
]
SECURITY_ALERT_CSV_COLUMNS = [
    "alert_id",
    "session_id",
    "alert_type",
    "reason_code",
    "student_code",
    "full_name",
    "class_name",
    "confidence_label",
    "confidence",
    "liveness_score",
    "gps_lat",
    "gps_lng",
    "captured_img",
    "dismissed",
    "dismissed_by",
    "dismissed_at",
    "note",
    "created_at",
]

router = APIRouter(prefix="/reports", tags=["Reports"])


def _safe_filename(value: str):
    return "".join(char if char.isalnum() or char in ("-", "_") else "_" for char in value).strip("_")


def _csv_value(value):
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _csv_response(rows, fieldnames, filename):
    stream = io.StringIO(newline="")
    writer = csv.DictWriter(stream, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({field: _csv_value(row.get(field)) for field in fieldnames})
    output = io.BytesIO(stream.getvalue().encode("utf-8-sig"))
    return StreamingResponse(
        output,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _parse_alert_note(note):
    if not note:
        return None, ""
    try:
        parsed = json.loads(note)
    except (TypeError, ValueError, json.JSONDecodeError):
        return None, note
    if not isinstance(parsed, dict):
        return None, note
    return parsed.get("reason_code"), note


def _confidence_label(alert_type):
    normalized = (alert_type or "").upper()
    if normalized == "FACE_UNCLEAR":
        return "Độ tin cậy phát hiện khuôn mặt"
    if normalized in {"UNKNOWN_FACE", "UNKNOWN"}:
        return "Độ tin cậy khớp danh tính"
    if normalized == "SPOOF":
        return "Độ tin cậy cảnh báo"
    return "Độ tin cậy"


def _attendance_csv_rows(session, report_rows, db: Session):
    record_ids = [row.get("record_id") for row in report_rows if row.get("record_id")]
    records = {}
    if record_ids:
        records = {
            record.id: record
            for record in db.query(Attendance).filter(Attendance.id.in_(record_ids)).all()
        }

    rows = []
    for row in report_rows:
        record = records.get(row.get("record_id"))
        rows.append(
            {
                "session_id": session.id,
                "class_name": row.get("class_name") or session.class_name,
                "student_code": row.get("student_code"),
                "full_name": row.get("full_name"),
                "attendance_status": row.get("status"),
                "check_in_at": record.check_in_at if record else row.get("check_in_at"),
                "check_out_at": record.check_out_at if record else row.get("check_out_at"),
                "check_in_conf": record.check_in_conf if record else row.get("check_in_conf"),
                "check_out_conf": record.check_out_conf if record else row.get("check_out_conf"),
                "liveness_passed": record.liveness_passed if record else None,
                "gps_lat": record.gps_lat if record else None,
                "gps_lng": record.gps_lng if record else None,
                "gps_accuracy": record.gps_accuracy if record else None,
                "distance_meters": record.distance_meters if record else None,
                "check_in_img": record.check_in_img if record else None,
                "scan_count": record.scan_count if record else None,
                "last_scan_at": record.last_scan_at if record else None,
                "note": record.note if record else row.get("note"),
                "created_at": record.created_at if record else None,
            }
        )
    return rows


def _security_alert_csv_rows(session_id: int, db: Session):
    alerts = (
        db.query(SecurityAlert, Student)
        .outerjoin(Student, SecurityAlert.student_id == Student.id)
        .filter(SecurityAlert.session_id == session_id)
        .order_by(SecurityAlert.created_at.asc(), SecurityAlert.id.asc())
        .all()
    )

    rows = []
    for alert, student in alerts:
        reason_code, note = _parse_alert_note(alert.note)
        rows.append(
            {
                "alert_id": alert.id,
                "session_id": alert.session_id,
                "alert_type": alert.alert_type,
                "reason_code": reason_code,
                "student_code": student.student_code if student else None,
                "full_name": student.full_name if student else None,
                "class_name": student.class_name if student else None,
                "confidence_label": _confidence_label(alert.alert_type),
                "confidence": alert.confidence,
                "liveness_score": alert.liveness_score,
                "gps_lat": alert.gps_lat,
                "gps_lng": alert.gps_lng,
                "captured_img": alert.captured_img,
                "dismissed": alert.dismissed,
                "dismissed_by": alert.dismissed_by,
                "dismissed_at": alert.dismissed_at,
                "note": note,
                "created_at": alert.created_at,
            }
        )
    return rows


def _new_pdf(stream):
    c = canvas.Canvas(stream, pagesize=A4)
    return c, PAGE_H - MARGIN


def _check_page(c, y, line_h=16, reserve=60):
    if y - line_h < reserve:
        c.showPage()
        c.setFont(_PDF_FONT, 10)
        return PAGE_H - MARGIN
    return y


def _draw_header(c, y, title, subtitle=""):
    c.setFont(_PDF_FONT_BOLD, 13)
    c.drawString(MARGIN, y, title)
    y -= 18
    if subtitle:
        c.setFont(_PDF_FONT, 10)
        c.drawString(MARGIN, y, subtitle)
        y -= 14
    c.setLineWidth(0.5)
    c.line(MARGIN, y, PAGE_W - MARGIN, y)
    y -= 12
    return y


def _draw_row(c, y, cols, widths, bold=False):
    font = _PDF_FONT_BOLD if bold else _PDF_FONT
    c.setFont(font, 9)
    x = MARGIN
    for text, width in zip(cols, widths):
        c.drawString(x + 2, y, str(text)[: int(width / 5.5)])
        x += width
    return y - 14


def _fmt_dt(iso):
    if not iso:
        return "-"
    try:
        from datetime import datetime

        value = datetime.fromisoformat(str(iso))
        return value.strftime("%d/%m %H:%M")
    except Exception:
        return str(iso)[:16]


def _metric_value(metrics, label, default=0):
    match = metrics.loc[metrics.iloc[:, 0] == label]
    if match.empty:
        return default
    try:
        return float(match.iloc[0, 1])
    except Exception:
        return default


def _format_percent(value):
    try:
        return f"{float(value):.1%}"
    except Exception:
        return "-"


def _format_confidence(value):
    if not isinstance(value, (int, float)):
        return "-"
    return f"{value * 100:.1f}%"


def _format_time_range(start_time, end_time):
    if not start_time and not end_time:
        return "Chưa đặt giờ"
    return f"{start_time or 'Chưa đặt'} - {end_time or 'Chưa đặt'}"


def _summary_export_frame(rows):
    if not rows:
        return pd.DataFrame(columns=list(SUMMARY_COLUMNS.values()))

    normalized = []
    for row in rows:
        item = dict(row)
        item["rate"] = _format_percent(item.get("rate"))
        item["warning"] = "Có" if item.get("warning") else "Không"
        normalized.append(item)
    return pd.DataFrame(normalized).rename(columns=SUMMARY_COLUMNS)


def _session_export_frame(rows):
    normalized = []
    for row in rows:
        item = dict(row)
        item["status"] = STATUS_LABELS.get(item.get("status"), item.get("status"))
        item["start_time"] = item.get("start_time") or "-"
        item["end_time"] = item.get("end_time") or "-"
        item["check_in_at"] = _fmt_dt(item.get("check_in_at"))
        item["check_out_at"] = _fmt_dt(item.get("check_out_at"))
        item["check_in_conf"] = _format_confidence(item.get("check_in_conf"))
        item["check_out_conf"] = _format_confidence(item.get("check_out_conf"))
        normalized.append(item)
    return pd.DataFrame(normalized).rename(columns=SESSION_COLUMNS)


def _read_evaluation_reports():
    if not EVALUATION_METRICS_CSV.exists() or not EVALUATION_DETAILS_CSV.exists():
        return None, None

    try:
        metrics = pd.read_csv(EVALUATION_METRICS_CSV)
        details = pd.read_csv(EVALUATION_DETAILS_CSV)
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail="Không đọc được báo cáo đánh giá mô hình. Vui lòng kiểm tra lại tệp báo cáo.",
        ) from exc

    return metrics, _normalize_evaluation_details(details)


def _normalize_evaluation_details(details):
    def clean_code(value):
        if value is None or pd.isna(value):
            return None
        text_value = str(value)
        if text_value.endswith(".0"):
            return text_value[:-2]
        return text_value

    details = details.copy()
    if "image_path" not in details:
        details["image_path"] = ""
    if "file_name" not in details:
        details["file_name"] = details["image_path"].apply(lambda value: Path(str(value)).name if value else "")
    if "sample_name" not in details:
        details["sample_name"] = details["file_name"].apply(lambda value: Path(str(value)).stem if value else "")
    if "actual_student_code" not in details:
        details["actual_student_code"] = details.get("true_label", "")
    if "sample_code" not in details:
        details["sample_code"] = details["actual_student_code"]
    if "expected_student_code" not in details:
        details["expected_student_code"] = details["actual_student_code"]
    if "mapping_source" not in details:
        details["mapping_source"] = ""
    if "predicted_student_code" not in details:
        details["predicted_student_code"] = ""
    if "status" not in details:
        details["status"] = ""
    if "confidence" not in details:
        details["confidence"] = None
    if "processing_time_ms" not in details:
        details["processing_time_ms"] = None
    if "result" not in details:
        details["result"] = ""

    details = details.where(pd.notnull(details), None)
    for column in ("actual_student_code", "sample_code", "expected_student_code", "predicted_student_code"):
        if column in details:
            details[column] = details[column].apply(clean_code)
    return details


def _evaluation_metric_summary(metrics, details):
    total_images = int(_metric_value(metrics, "Total images", len(details)))
    tp = int(_metric_value(metrics, "TP", 0))
    tn = int(_metric_value(metrics, "TN", 0))
    fp = int(_metric_value(metrics, "FP", 0))
    fn = int(_metric_value(metrics, "FN", 0))

    result_series = details["result"].fillna("").astype(str).str.upper() if "result" in details else pd.Series(dtype=str)
    status_series = details["status"].fillna("").astype(str) if "status" in details else pd.Series(dtype=str)
    confidence_values = pd.to_numeric(details.get("confidence"), errors="coerce")
    confidence_values = confidence_values[confidence_values >= 0]
    processing_values = pd.to_numeric(details.get("processing_time_ms"), errors="coerce").dropna()

    return {
        "has_data": total_images > 0 and not details.empty,
        "source": "evaluation_csv",
        "sample_count": total_images,
        "total_images": total_images,
        "recognized_correct": int((result_series == "TP").sum()),
        "recognized_wrong": int((result_series == "FP").sum()),
        "not_recognized": int(status_series.isin(["unknown", "no_face", "multiple_faces"]).sum()),
        "tp": tp,
        "tn": tn,
        "fp": fp,
        "fn": fn,
        "accuracy": _metric_value(metrics, "Accuracy", 0),
        "precision": _metric_value(metrics, "Precision", 0),
        "recall": _metric_value(metrics, "Recall", 0),
        "f1_score": _metric_value(metrics, "F1-score", 0),
        "far": _metric_value(metrics, "FAR", 0),
        "frr": _metric_value(metrics, "FRR", 0),
        "average_confidence": round(float(confidence_values.mean()), 4) if not confidence_values.empty else 0,
        "average_processing_time_ms": round(float(processing_values.mean()), 2) if not processing_values.empty else None,
        "detail_count": int(len(details)),
    }


def _evaluation_details_records(details):
    columns = [
        "file_name",
        "sample_name",
        "actual_student_code",
        "sample_code",
        "expected_student_code",
        "predicted_student_code",
        "status",
        "confidence",
        "processing_time_ms",
        "result",
        "mapping_source",
        "dataset_type",
        "image_path",
    ]
    for column in columns:
        if column not in details:
            details[column] = None
    return details[columns].to_dict(orient="records")


def _format_excel_sheet(worksheet):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(name="Arial", bold=True, color="111827")
    body_font = Font(name="Arial", color="111827")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_length + 2, 10), 34)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    worksheet.freeze_panes = "A2"


@router.get("/dashboard/stats")
def get_dashboard_stats(current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    return report_service.get_dashboard_stats_for_user(db, current_user)


@router.get("/model-evaluation/stats")
def get_model_evaluation_stats(_current_user=Depends(require_admin)):
    metrics, details = _read_evaluation_reports()
    if metrics is None or details is None:
        return {
            "has_data": False,
            "message": (
                "Chưa có dữ liệu đánh giá mô hình. "
                "Vui lòng chạy evaluate_recognition.py với ảnh kiểm thử thật trước khi lấy số liệu báo cáo."
            ),
        }

    summary = _evaluation_metric_summary(metrics, details)
    if not summary["has_data"]:
        return {
            "has_data": False,
            "message": "Tệp đánh giá đã tồn tại nhưng chưa có mẫu kiểm thử hợp lệ.",
        }
    return summary


@router.get("/model-evaluation/details")
def get_model_evaluation_details(_current_user=Depends(require_admin)):
    metrics, details = _read_evaluation_reports()
    if metrics is None or details is None or details.empty:
        return {"has_data": False, "items": []}

    return {
        "has_data": True,
        "summary": _evaluation_metric_summary(metrics, details),
        "items": _evaluation_details_records(details),
    }


@router.get("/export/model-evaluation/csv")
def export_model_evaluation_csv(_current_user=Depends(require_admin)):
    metrics, details = _read_evaluation_reports()
    if metrics is None or details is None or details.empty:
        raise HTTPException(status_code=404, detail="Không tìm thấy dữ liệu đánh giá mô hình.")

    stream = io.StringIO()
    pd.DataFrame(_evaluation_details_records(details)).to_csv(stream, index=False, encoding="utf-8-sig")
    output = io.BytesIO(stream.getvalue().encode("utf-8-sig"))
    return StreamingResponse(
        output,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=model_evaluation_details.csv"},
    )


@router.get("/export/model-evaluation/excel")
def export_model_evaluation_excel(_current_user=Depends(require_admin)):
    metrics, details = _read_evaluation_reports()
    if metrics is None or details is None or details.empty:
        raise HTTPException(status_code=404, detail="Không tìm thấy dữ liệu đánh giá mô hình.")

    summary = _evaluation_metric_summary(metrics, details)
    summary_rows = [
        {"metric": "Tổng ảnh kiểm thử", "value": summary["total_images"]},
        {"metric": "Nhận diện đúng", "value": summary["recognized_correct"]},
        {"metric": "Nhận diện sai", "value": summary["recognized_wrong"]},
        {"metric": "Không nhận diện được", "value": summary["not_recognized"]},
        {"metric": "TP", "value": summary["tp"]},
        {"metric": "FP", "value": summary["fp"]},
        {"metric": "FN", "value": summary["fn"]},
        {"metric": "TN", "value": summary["tn"]},
        {"metric": "Độ chính xác", "value": summary["accuracy"]},
        {"metric": "Độ chính xác dự đoán", "value": summary["precision"]},
        {"metric": "Độ bao phủ", "value": summary["recall"]},
        {"metric": "Điểm F1", "value": summary["f1_score"]},
        {"metric": "Tỷ lệ chấp nhận sai (FAR)", "value": summary["far"]},
        {"metric": "Tỷ lệ từ chối sai (FRR)", "value": summary["frr"]},
        {"metric": "Độ tin cậy trung bình", "value": summary["average_confidence"]},
        {"metric": "Thời gian xử lý trung bình (ms)", "value": summary["average_processing_time_ms"]},
    ]

    stream = io.BytesIO()
    with pd.ExcelWriter(stream, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name="Chỉ số")
        pd.DataFrame(_evaluation_details_records(details)).to_excel(writer, index=False, sheet_name="Chi tiết")
        _format_excel_sheet(writer.sheets["Chỉ số"])
        _format_excel_sheet(writer.sheets["Chi tiết"])
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=model_evaluation.xlsx"},
    )


@router.get("/summary/{class_name}")
def get_summary_by_class(class_name: str, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    return report_service.build_class_summary_for_user(class_name, db, current_user)


@router.get("/warnings/{class_name}")
def get_warnings_by_class(class_name: str, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    summary = report_service.build_class_summary_for_user(class_name, db, current_user)
    return [item for item in summary if item["warning"]]


@router.get("/session/{session_id}")
def get_session_report(session_id: int, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    _session, report_rows = report_service.build_session_report_for_user(session_id, db, current_user)
    return report_rows


@router.get("/export/excel/{class_name}")
def export_excel(class_name: str, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    summary = report_service.build_class_summary_for_user(class_name, db, current_user)
    if not summary:
        raise HTTPException(status_code=404, detail="Không tìm thấy dữ liệu trong lớp này.")

    warnings = [item for item in summary if item["warning"]]
    stream = io.BytesIO()
    with pd.ExcelWriter(stream, engine="openpyxl") as writer:
        _summary_export_frame(summary).to_excel(writer, index=False, sheet_name="Tong hop")
        _summary_export_frame(warnings).to_excel(writer, index=False, sheet_name="Canh bao")
        _format_excel_sheet(writer.sheets["Tong hop"])
        _format_excel_sheet(writer.sheets["Canh bao"])
    stream.seek(0)

    filename = _safe_filename(f"attendance_{class_name}") or "attendance_report"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


@router.get("/export/pdf/{class_name}")
def export_pdf(class_name: str, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    summary = report_service.build_class_summary_for_user(class_name, db, current_user)
    if not summary:
        raise HTTPException(status_code=404, detail="Không tìm thấy dữ liệu trong lớp này.")

    stream = io.BytesIO()
    c, y = _new_pdf(stream)

    title = f"Báo cáo chuyên cần - Lớp {class_name}"
    subtitle = f"Tổng số sinh viên: {len(summary)}  |  Tổng buổi học: {summary[0]['total_sessions']}"
    y = _draw_header(c, y, title, subtitle)

    headers = ["Mã SV", "Họ tên", "Có mặt", "Trễ", "Vắng mặt", "Tổng", "Tỷ lệ", "Cảnh báo"]
    widths = [80, 220, 55, 45, 50, 50, 60, 60]
    y = _draw_row(c, y, headers, widths, bold=True)
    c.setLineWidth(0.3)
    c.line(MARGIN, y + 10, PAGE_W - MARGIN, y + 10)

    for item in summary:
        y = _check_page(c, y)
        cols = [
            item["student_code"],
            item["full_name"],
            item["present"],
            item["late"],
            item["absent"],
            item["total_sessions"],
            f"{item['rate']:.1%}",
            "!" if item["warning"] else "",
        ]
        y = _draw_row(c, y, cols, widths)

    c.save()
    stream.seek(0)

    filename = _safe_filename(f"attendance_{class_name}") or "attendance_report"
    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}.pdf"},
    )


@router.get("/export/excel/warnings/{class_name}")
def export_warning_excel(class_name: str, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    summary = report_service.build_class_summary_for_user(class_name, db, current_user)
    if not summary:
        raise HTTPException(status_code=404, detail="Không tìm thấy dữ liệu của lớp này.")

    warnings = [item for item in summary if item["warning"]]
    stream = io.BytesIO()
    with pd.ExcelWriter(stream, engine="openpyxl") as writer:
        _summary_export_frame(warnings).to_excel(writer, index=False, sheet_name="Canh bao")
        _format_excel_sheet(writer.sheets["Canh bao"])
    stream.seek(0)

    filename = _safe_filename(f"attendance_warnings_{class_name}") or "attendance_warnings"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


@router.get("/export/csv/session/{session_id}")
def export_session_csv(session_id: int, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    session, rows = report_service.build_session_report_for_user(session_id, db, current_user)
    csv_rows = _attendance_csv_rows(session, rows, db)
    filename = _safe_filename(
        f"attendance_session_{session.id}_{session.class_name}_{session.session_date}"
    ) or f"attendance_session_{session.id}"
    return _csv_response(csv_rows, ATTENDANCE_CSV_COLUMNS, f"{filename}.csv")


@router.get("/export/csv/session/{session_id}/alerts")
def export_session_alerts_csv(session_id: int, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    session, _rows = report_service.build_session_report_for_user(session_id, db, current_user)
    csv_rows = _security_alert_csv_rows(session_id, db)
    filename = _safe_filename(
        f"security_alerts_session_{session.id}_{session.class_name}_{session.session_date}"
    ) or f"security_alerts_session_{session.id}"
    return _csv_response(csv_rows, SECURITY_ALERT_CSV_COLUMNS, f"{filename}.csv")


@router.get("/export/excel/session/{session_id}")
def export_session_excel(session_id: int, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    session, rows = report_service.build_session_report_for_user(session_id, db, current_user)
    if not rows:
        raise HTTPException(status_code=404, detail="Không tìm thấy dữ liệu của buổi học này.")

    stream = io.BytesIO()
    with pd.ExcelWriter(stream, engine="openpyxl") as writer:
        _session_export_frame(rows).to_excel(writer, index=False, sheet_name="Bao cao buoi")
        _format_excel_sheet(writer.sheets["Bao cao buoi"])
    stream.seek(0)

    filename = _safe_filename(
        f"attendance_session_{session.id}_{session.class_name}_{session.session_date}"
    ) or f"attendance_session_{session.id}"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


@router.get("/export/pdf/session/{session_id}")
def export_session_pdf(session_id: int, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    session, rows = report_service.build_session_report_for_user(session_id, db, current_user)
    if not rows:
        raise HTTPException(status_code=404, detail="Không tìm thấy dữ liệu của buổi học này.")

    stream = io.BytesIO()
    c, y = _new_pdf(stream)

    title = f"Báo cáo buổi học #{session.id} - {session.class_name} - {session.subject}"
    subtitle = (
        f"Ngày: {session.session_date}  |  "
        f"Giờ: {_format_time_range(session.start_time, session.end_time)}  |  "
        f"Sĩ số: {len(rows)}"
    )
    y = _draw_header(c, y, title, subtitle)

    headers = ["Mã SV", "Họ tên", "Trạng thái", "Vào lớp", "Ra về", "Độ tin cậy"]
    widths = [80, 200, 90, 110, 110, 65]
    y = _draw_row(c, y, headers, widths, bold=True)
    c.setLineWidth(0.3)
    c.line(MARGIN, y + 10, PAGE_W - MARGIN, y + 10)

    present_count = sum(1 for row in rows if row["status"] in ("present", "late", "left_early"))
    absent_count = sum(1 for row in rows if row["status"] == "absent")

    for item in rows:
        y = _check_page(c, y)
        confidence = item.get("check_in_conf")
        confidence_text = f"{confidence * 100:.1f}%" if isinstance(confidence, float) else "-"
        cols = [
            item["student_code"],
            item["full_name"],
            STATUS_LABELS.get(item["status"], item["status"]),
            _fmt_dt(item.get("check_in_at")),
            _fmt_dt(item.get("check_out_at")),
            confidence_text,
        ]
        y = _draw_row(c, y, cols, widths)

    y = _check_page(c, y, reserve=40)
    y -= 8
    c.setLineWidth(0.5)
    c.line(MARGIN, y, PAGE_W - MARGIN, y)
    y -= 14
    c.setFont(_PDF_FONT_BOLD, 9)
    c.drawString(MARGIN, y, f"Có mặt: {present_count}   Vắng: {absent_count}   Tổng: {len(rows)}")

    c.save()
    stream.seek(0)

    filename = _safe_filename(
        f"attendance_session_{session.id}_{session.class_name}_{session.session_date}"
    ) or f"attendance_session_{session.id}"
    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}.pdf"},
    )
